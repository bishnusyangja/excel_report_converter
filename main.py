from datetime import datetime

from openpyxl import load_workbook
from openpyxl import Workbook

from settings import INPUT_FILE
from settings import INPUT_SHEET
from settings import OUTPUT_FILE
from settings import OUTPUT_SHEET


def is_last_row(row):
    site_id = row[0].value
    if site_id and isinstance(site_id, str):
        site_tokens = site_id.split()
        try:
            int(site_tokens[1])
            return site_id
        except Exception as e:
            print("Excep: ", e)
            return
    return


def format_row_item(row_item, row_id, skip_key="Site ID"):
    formatted_row = {}
    row_len = len(row_item)
    for j, item in enumerate(row_item[1]):
        if item not in formatted_row and item != skip_key:
            formatted_row[item] = []
        if j == 0: # discarding first column
            continue
        try:
            key_item = row_item[1][j]
        except Exception as e:
            continue
        if key_item in formatted_row:
            temp = []
            for i in range(row_len):
                if i == 1:
                    continue
                try:
                    temp.append(row_item[i][j])
                except Exception as e:
                    print("temp exception : ", row_id, "i=",i, "j=", j, key_item, len(row_item[i]), row_item[1][j])
            formatted_row[key_item].append(temp)
    return formatted_row


def write_data_to_excel_file(data, header):
    wb = Workbook()
    sheet = wb.active
    sheet.title = OUTPUT_SHEET
    row_num = 1
    for i, heading in enumerate(header):
        sheet.cell(row=row_num, column=i+1, value=heading)
    row_num += 1

    for site_id, content in data.items():
        for date_value, values in content.items():
            sheet.cell(row=row_num, column=3, value=site_id)
            sheet.cell(row=row_num, column=1, value=date_value.day)
            sheet.cell(row=row_num, column=2, value=date_value.strftime("%Y/%m/%d"))
            for item in values:
                heading = item[0]
                value = item[1]
                col = header.index(heading)+1
                sheet.cell(row=row_num, column=col, value=value)
            row_num += 1
    wb.save(OUTPUT_FILE)


def get_header_from_first_row(row):
    header_unique = list(set(row))
    header_filtered = [item for item in header_unique if not isinstance(item, datetime)]
    header = ['Day of Month', 'Date', 'Site ID']
    header.extend(header_filtered)
    return header


def read_data_from_sheet(sheet):
    data = {}
    row_item = []
    header = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=0, min_col=0, max_row=sheet.max_row, max_col=sheet.max_column)):
        temp = []
        if row_num == 0:
            continue
        row_id = is_last_row(row)
        for col_i, cell in enumerate(row):
            if cell.value is not None:
                temp.append(cell.value)
        row_item.append(temp)
        if row_id:
            formatted_row = format_row_item(row_item, row_id)
            if formatted_row:
                data[row_id] = formatted_row
            if not header:
                header = get_header_from_first_row(row_item[0])
            row_item = []
    return data, header


def run():
    wb = load_workbook(INPUT_FILE)
    sheet = wb[INPUT_SHEET]
    data, header = read_data_from_sheet(sheet)
    write_data_to_excel_file(data, header)


if __name__ == '__main__':
    run()
