import unittest
from datetime import datetime

from openpyxl import Workbook
from main import read_data_from_sheet, is_last_row, format_row_item, write_data_to_excel_sheet, \
    get_header_from_first_row


class TestMain(unittest.TestCase):

    def write_to_sheet(self, sheet, data):
        for i, row in enumerate(data):
            for j, item in enumerate(row):
                sheet.cell(row=i+1, column=j+1, value=item)
        return sheet

    def test_is_last_row_with_invalid(self):
        site_id = "site abc"
        result = is_last_row(site_id)
        self.assertEqual(result, None)

    def test_is_last_row_with_valid(self):
        site_id = "site 14"
        result = is_last_row(site_id)
        self.assertEqual(result, site_id)

    def test_format_row_item(self):
        row_item = [['Page Views', 'Page Views', 'Total Time Spent', 'Total Time spent', 'Visitors'],
                    ['2022-01-01', '2022-01-02', '2022-01-01', '2022-01-02', '2022-01-04'],
                    [4, 8, 9, 12, 14]
                    ]
        row_id = 'site 100'
        result = format_row_item(row_item, row_id,)
        expected = {'2022-01-01': [['Total Time Spent', 9]], '2022-01-02': [['Page Views', 8],
                            ['Total Time spent', 12]], '2022-01-04': [['Visitors', 14]]}
        self.assertEqual(result, expected)

    def test_get_header_from_first_row(self):
        row = ['Page Views', 'Page Views', 'Random', 'Random', 'Sample', 'Page Views', 'Random', 'Random']
        base_header = ['Day of Month', 'Date', 'Site ID']
        result = get_header_from_first_row(row)
        self.assertIn('Day of Month', result)
        self.assertIn('Date', result)
        self.assertIn('Site ID', result)
        self.assertIn('Page Views', result)
        self.assertIn('Random', result)
        self.assertIn('Sample', result)
        self.assertEqual(len(result), 6)

    def test_read_data_from_sheet(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Test Sheet"
        data = [['', '', ''], ['abcd', 'jpt', 'random'], ['site 2', 'xyz', 'root'], ['random', 'sample', 'pqr']]
        sheet = self.write_to_sheet(sheet, data)
        data, header = read_data_from_sheet(sheet)
        self.assertGreater(len(header), 3)
        self.assertEqual(type(data), dict)

    def test_write_data_to_excel_sheet(self):
        wb = Workbook()
        sheet = wb.active
        header = ['Day', 'Page Views', 'Visitors']
        data = {'site 5':  {datetime(2022, 2, 1): [['Page Views', 20], ['Visitors', 10]], datetime(2022, 2, 2):
            [['Page Views', 50], ['Visitors', 10]]}}
        result = write_data_to_excel_sheet(data, header, sheet)
        self.assertEqual(type(result), type(sheet))


if __name__ == "__main__":
    unittest.main()